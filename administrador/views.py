from django.shortcuts import render, redirect

from registro.models import Registro

from registro.forms import RegistroForm



from django.contrib.auth import login, authenticate
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.forms import UserCreationForm
from django.contrib import messages
from django.contrib.auth import logout
from .models import Asistencia
from django.utils import timezone

import openpyxl
import pandas as pd

from .forms import EntradaForm, SalidaForm

from datetime import timedelta

from django.db import IntegrityError
from django.utils.timezone import make_naive

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from django.http import HttpResponse
from django.utils import timezone

from django.db.models import Count, Q
from django.core.paginator import Paginator, EmptyPage
from django.http import HttpResponseRedirect
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required, user_passes_test

from django.db.models.functions import TruncDate

from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse


# Verifica que el usuario sea administrador
def is_admin(user):
    return user.groups.filter(name='Administrador').exists()

# Verifica que el usuario sea de asistencia
def is_asistencia(user):
    return user.groups.filter(name='Asistencia').exists()


@user_passes_test(is_admin)
def index(request):
    # CANTIDADES
    inscritos_count = Registro.objects.all().count()
    inscritosValidados_count = Registro.objects.filter(validado=True).count() 


    
    # Obtener los conteos de entradas y salidas por fecha
    asistencias = Asistencia.objects.values('fecha') \
                                    .annotate(
                                        entradas_count=Count('entrada', filter=~Q(entrada=None)),
                                        salidas_count=Count('salida', filter=~Q(salida=None))
                                    ) \
                                    .order_by('-fecha')
    
    # roles
    is_admin = request.user.groups.filter(name='Administrador').exists()
    is_asistencia = request.user.groups.filter(name='Asistencia').exists()

    # registros lugares
    ubicacion_a = Asistencia.objects.filter(ubicacion='A').values('fecha') \
                                    .annotate(
                                        entradas_count=Count('entrada', filter=~Q(entrada=None)),
                                        salidas_count=Count('salida', filter=~Q(salida=None))
                                    ) \
                                    .order_by('-fecha')
    
    ubicacion_b = Asistencia.objects.filter(ubicacion='B').values('fecha') \
                                    .annotate(
                                        entradas_count=Count('entrada', filter=~Q(entrada=None)),
                                        salidas_count=Count('salida', filter=~Q(salida=None))
                                    ) \
                                    .order_by('-fecha')
    
    ubicacion_c = Asistencia.objects.filter(ubicacion='C').values('fecha') \
                                    .annotate(
                                        entradas_count=Count('entrada', filter=~Q(entrada=None)),
                                        salidas_count=Count('salida', filter=~Q(salida=None))
                                    ) \
                                    .order_by('-fecha')
    
    ubicacion_d = Asistencia.objects.filter(ubicacion='D').values('fecha') \
                                    .annotate(
                                        entradas_count=Count('entrada', filter=~Q(entrada=None)),
                                        salidas_count=Count('salida', filter=~Q(salida=None))
                                    ) \
                                    .order_by('-fecha')
    
    # Registro por fechas
    registro_fechas = Registro.objects.filter(fecha_registro__isnull=False)

    fechas_contadas = {}

    for registro in registro_fechas:
        fecha = registro.fecha_registro.date()  # `.date()` obtiene solo la parte de la fecha
        if fecha in fechas_contadas:
            fechas_contadas[fecha] += 1
        else:
            fechas_contadas[fecha] = 1

    registroPorFechas = []

    for fecha, total in fechas_contadas.items():
        registroPorFechas.append({'fecha': fecha.strftime('%Y-%m-%d'), 'total': total})

    
    # Participantes y estudiantes
    participantes_count = Registro.objects.filter(tipo_participante='participante').count()
    estudiantes_count = Registro.objects.filter(tipo_participante='estudiante (pregrado)').count()


    # Renderizar
    return render(request, 'administrador/index-inscritos.html', {
        'inscritos_count': inscritos_count,
        'inscritosValidados_count': inscritosValidados_count,
        'asistencias': asistencias,
        'ubicacion_a': ubicacion_a,
        'ubicacion_b': ubicacion_b,
        'ubicacion_c': ubicacion_c,
        'ubicacion_d': ubicacion_d,
        'is_admin': is_admin,
        'is_asistencia': is_asistencia,
        'registroPorFechas': registroPorFechas,
        'participantes_count': participantes_count,
        'estudiantes_count': estudiantes_count
    })



@user_passes_test(is_admin)
def lista_inscritos(request):
    # Obtén todos los registros de inscritos ordenados por fecha
    inscritos = Registro.objects.all().order_by('-fecha_registro')

    # Paginación
    paginator = Paginator(inscritos, 10)  # Muestra 10 inscritos por página
    page_number = request.GET.get('page')  # Obtén el número de página desde la URL

    try:
        page_obj = paginator.page(page_number)  # Obtén los registros correspondientes a esa página
    except EmptyPage:
        page_obj = paginator.page(1)  # Si la página no existe o es vacía, muestra la primera página
    except:
        page_obj = paginator.page(paginator.num_pages)  # Si la página no existe, muestra la última página


    # roles
    is_admin = request.user.groups.filter(name='Administrador').exists()
    is_asistencia = request.user.groups.filter(name='Asistencia').exists()

    return render(request, 'administrador/lista-inscritos.html', {'page_obj': page_obj, 'is_admin': is_admin, 'is_asistencia': is_asistencia})



def validar_inscrito(request, id):
    if request.method == 'POST':
        try:
            inscrito = Registro.objects.get(id=id)
            inscrito.validado = not inscrito.validado
            inscrito.save()
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'status': 'success',
                    'new_state': inscrito.validado
                })
                
            referer = request.META.get('HTTP_REFERER')
            return HttpResponseRedirect(referer if referer else 'lista-inscritos')
            
        except Registro.DoesNotExist:
            return JsonResponse({'status': 'error'}, status=404)




def excel_inscritos_validados(request):
    # Filtra los registros validados
    inscritos = Registro.objects.filter(validado=1)

    # Organiza los datos en un diccionario
    data = {
        'DNI': [inscrito.dni for inscrito in inscritos],
        'Nombre': [inscrito.nombres for inscrito in inscritos],
        'Apellido Paterno': [inscrito.apellido_paterno for inscrito in inscritos],
        'Apellido Materno': [inscrito.apellido_materno for inscrito in inscritos],
        'Email': [inscrito.email for inscrito in inscritos],
        'Celular': [inscrito.celular for inscrito in inscritos],
        'Fecha de registro': [
            # Accede a los atributos correctamente usando la notación de puntos
            make_naive(inscrito.fecha_registro).strftime('%Y-%m-%d %H:%M:%S') if inscrito.fecha_registro else None
            for inscrito in inscritos
        ],
        # 'Validado': [inscrito.validado for inscrito in inscritos],
    }

    # Crea un DataFrame de pandas
    df = pd.DataFrame(data)

    # Crea la respuesta HTTP con el tipo de contenido adecuado
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=inscritos_validados.xlsx'

    # df.to_excel(response, index=False)

    with pd.ExcelWriter(response, engine='openpyxl') as writer:

        df.to_excel(writer, index=False, sheet_name='Inscritos', startrow=4)  # Iniciar en 4

        worksheet = writer.sheets['Inscritos']
        
        worksheet.merge_cells('A1:G1')         

        worksheet['A2'] = 'CONGRESO CIMAC 2025'
        worksheet['A3'] = 'LISTA DE PERSONAS INSCRITAS'
        
     
     
        worksheet['A2'].font = Font(bold=True, size=16, color="000000")  
        worksheet['A2'].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") 
        
        worksheet.merge_cells('A2:G2')
        worksheet.row_dimensions[2].height = 30
        worksheet['A2'].alignment = Alignment(horizontal='center', vertical='center')
        
        
        worksheet['A3'].font = Font(bold=True, size=14, color="000000") 
        worksheet['A3'].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") 
        
        worksheet.merge_cells('A3:G3') 
        worksheet.row_dimensions[3].height = 20
        worksheet['A3'].alignment = Alignment(horizontal='center', vertical='center')
        
        worksheet.merge_cells('A4:G4')         
                
        def ajustar_ancho_columnas(worksheet, dataframe):
            
            border = Border(
                left=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000"),
                top=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000")
            )
            
            for i, col in enumerate(dataframe.columns, 1):
                max_len = max(dataframe[col].astype(str).map(len).max(), len(col))  # Considera el largo del encabezado
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max_len + 2  # Añadir un margen
                
            # Agregar bordes a todas las celdas
            for row in worksheet.iter_rows(min_row=6, min_col=1, max_row=worksheet.max_row, max_col=len(dataframe.columns)):
                for cell in row:
                    cell.border = border

        # Ajustar ancho de columnas
        ajustar_ancho_columnas(worksheet, df)

    return response




def entrada_inscritos(request):
    # roles
    is_admin = request.user.groups.filter(name='Administrador').exists()
    is_asistencia = request.user.groups.filter(name='Asistencia').exists()

    
    inicio_dia = timezone.make_aware(timezone.datetime.combine(timezone.now().date(), timezone.datetime.min.time()))
    fin_dia = inicio_dia + timedelta(days=1) - timedelta(seconds=1)
    
    if request.method == 'POST':
        form = EntradaForm(request.POST)
        if form.is_valid():
            dni = form.cleaned_data['dni']
            ubicacion = form.cleaned_data['ubicacion']

            try:
                # Buscar al usuario con el DNI proporcionado
                usuario = Registro.objects.get(dni=dni)

                # Verificar si ya existe un registro de asistencia para este usuario y la fecha actual
                asistencia = Asistencia.objects.filter(usuario=usuario, fecha=timezone.now().date()).first()
                if asistencia:
                    form.add_error('dni', 'DNI ya registrado.')
                else:
                    hora_ajustada = timezone.localtime(timezone.now()) - timedelta(hours=5)
                    # Intentamos crear el registro
                    Asistencia.objects.create(
                        usuario=usuario,
                        entrada=hora_ajustada,
                        fecha=timezone.now().date(),
                        ubicacion=ubicacion
                    )
            
            except Registro.DoesNotExist:
                form.add_error('dni', 'DNI no existe.')

            # Redirigir o volver a renderizar la página después de procesar el formulario
            return render(request, 'administrador/entrada-inscritos.html', {
                'form': form,
                'asistenciaEntrada': Asistencia.objects.all().order_by('-entrada'),
                'asistenciaEntradaHoyCount': Asistencia.objects.filter(entrada__gte=inicio_dia, entrada__lt=fin_dia).count(),
                'is_admin': is_admin,
                'is_asistencia': is_asistencia,
                'ubicacion_seleccionada': form.cleaned_data.get('ubicacion', '')  # Mantén el valor de la ubicación seleccionada
            })
        
        else:
            form.add_error('dni', 'Seleccione una ubicación.')
    
    else:
        form = EntradaForm()

    return render(request, 'administrador/entrada-inscritos.html', {
        'form': form,
        'asistenciaEntrada': Asistencia.objects.all().order_by('-entrada'),
        'asistenciaEntradaHoyCount': Asistencia.objects.filter(entrada__gte=inicio_dia, entrada__lt=fin_dia).count(),
        'is_admin': is_admin,
        'is_asistencia': is_asistencia,
        'ubicacion_seleccionada': form.cleaned_data.get('ubicacion', '') if form.is_bound and form.is_valid() else ''
    })




def salida_inscritos(request):
    # roles
    is_admin = request.user.groups.filter(name='Administrador').exists()
    is_asistencia = request.user.groups.filter(name='Asistencia').exists()

    
    inicio_dia = timezone.make_aware(timezone.datetime.combine(timezone.now().date(), timezone.datetime.min.time()))
    fin_dia = inicio_dia + timedelta(days=1) - timedelta(seconds=1)
    
    if request.method == 'POST':
        form = SalidaForm(request.POST)
        if form.is_valid():
            dni = form.cleaned_data['dni']
            ubicacion = form.cleaned_data['ubicacion']

            try:
                # Buscar al usuario con el DNI proporcionado
                usuario = Registro.objects.get(dni=dni)

                asistencia = Asistencia.objects.filter(usuario=usuario, fecha=timezone.now().date()).first()
                
                if asistencia:
                    if asistencia.entrada is not None and asistencia.salida is None:
                        # Si ya existe un registro de entrada y no tiene salida, actualizamos la salida
                        hora_ajustada = timezone.localtime(timezone.now()) - timedelta(hours=5)
                        asistencia.salida = hora_ajustada
                        asistencia.save()
                    else:
                        # Si no tiene entrada o ya tiene salida, mostramos un error
                        form.add_error('dni', 'DNI sin entrada hoy.')
                else:
                    # Si no existe el registro de asistencia, mostramos un error
                    form.add_error('dni', 'DNI sin entrada hoy.')
            
            except Registro.DoesNotExist:
                form.add_error('dni', 'DNI no existe.')

            # Redirigir o volver a renderizar la página después de procesar el formulario
            return render(request, 'administrador/salida-inscritos.html', {
                'form': form,
                'asistenciaSalida': Asistencia.objects.filter(salida__isnull=False).order_by('-salida'),
                'asistenciaSalidaHoyCount': Asistencia.objects.filter(salida__gte=inicio_dia, entrada__lt=fin_dia).count(), 
                'is_admin': is_admin,
                'is_asistencia': is_asistencia,
                'ubicacion_seleccionada': form.cleaned_data.get('ubicacion', '') 
            })
        
        else:
            form.add_error('dni', 'Seleccione una ubicación.')
    
    else:
        form = SalidaForm()

    return render(request, 'administrador/salida-inscritos.html', {
        'form': form,
        'asistenciaSalida': Asistencia.objects.filter(salida__isnull=False).order_by('-salida'),
        'asistenciaSalidaHoyCount': Asistencia.objects.filter(salida__gte=inicio_dia, entrada__lt=fin_dia).count(), 
        'is_admin': is_admin,
        'is_asistencia': is_asistencia,
        'ubicacion_seleccionada': form.cleaned_data.get('ubicacion', '') if form.is_bound and form.is_valid() else ''
    })




def ver_pdf(request):
    context = {'pdf_url': '/media/pdfs/'}
    return render(request, 'show_pdf.html', context)



@login_required
def eliminar_inscrito(request, id):
    inscrito = get_object_or_404(Registro, id=id)
    if request.method == 'POST':
        inscrito.delete()
        return redirect('lista-inscritos')
    # Podrías agregar una página de confirmación si quieres
    return redirect('lista-inscritos')



@login_required
def editar_imagen(request, registro_id):
    registro = get_object_or_404(Registro, pk=registro_id)

    if request.method == 'POST':
        form = RegistroForm(request.POST, request.FILES, instance=registro)
        # Limitar solo el campo voucher_pago
        for field in list(form.fields):
            if field != 'voucher_pago':
                form.fields.pop(field)

        if form.is_valid():
            form.save()
            return redirect('lista-inscritos')
    else:
        form = RegistroForm(instance=registro)
        # Limitar solo el campo voucher_pago
        for field in list(form.fields):
            if field != 'voucher_pago':
                form.fields.pop(field)

    return render(request, 'editar_voucher.html', {'form': form, 'registro': registro})



# CREAR Y LOGIN DEL ADMINISTRADOR

# def login_view(request):
#     if request.method == 'POST':
#         form = AuthenticationForm(request, data=request.POST)
#         if form.is_valid():
#             # Autenticar al usuario
#             username = form.cleaned_data.get('username')
#             password = form.cleaned_data.get('password')
#             user = authenticate(request, username=username, password=password)
#             if user is not None:
#                 login(request, user)
#                 return redirect('index-inscritos')  # Redirige a la página de inicio
#             else:
#                 messages.error(request, 'Usuario o contraseña incorrectos.')
#         else:
#             messages.error(request, 'Usuario y/o contraseña incorrectos.')
#     else:
#         form = AuthenticationForm()

#     return render(request, 'administrador/login-administrador.html', {'form': form})



def login_view(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            # Obtiene el username y password
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            
            # Autenticar al usuario
            user = authenticate(request, username=username, password=password)
            
            if user is not None:
                # Verifica si el usuario está activo
                if user.is_active:
                    login(request, user)
                    
                    # Verifica el grupo al que pertenece el usuario
                    if user.groups.filter(name='Administrador').exists():
                        return redirect('index-inscritos')  
                    elif user.groups.filter(name='Asistencia').exists():
                        return redirect('entrada-inscritos')
                    else:
                        messages.error(request, 'El usuario no tiene acceso a ninguna de las áreas disponibles.')
                        return render(request, 'administrador/login-administrador.html', {'form': form})
                else:
                    messages.error(request, 'El usuario está inactivo. Contacta al administrador.')
            else:
                messages.error(request, 'Usuario o contraseña incorrectos.')
        else:
            messages.error(request, 'Formulario inválido. Por favor, revisa tus credenciales.')
    else:
        form = AuthenticationForm()

    return render(request, 'administrador/login-administrador.html', {'form': form})


def logout_view(request):
    logout(request)  # Cierra la sesión del usuario
    return redirect('login')  # Redirige a la página de login o cualquier otra página pública



# def register_view(request):
#     if request.method == 'POST':
#         form = UserCreationForm(request.POST)
#         if form.is_valid():
#             form.save()  # Guarda el nuevo usuario
#             messages.success(request, '¡Cuenta creada exitosamente! Ahora puedes iniciar sesión.')
#             return redirect('login')  # Redirige al login después de crear la cuenta
#         else:
#             messages.error(request, 'Por favor, corrige los errores.')
#     else:
#         form = UserCreationForm()

#     return render(request, 'administrador/crear-administrador.html', {'form': form})


# roles y permisos
from django.contrib.auth.models import Group
from .forms import CustomUserCreationForm

def register_view(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            # Guarda el nuevo usuario
            user = form.save()

            # Asignar grupo según el rol elegido
            role = form.cleaned_data['role']

            if role == 'administrador':
                # Asignar al grupo 'Administrador'
                admin_group = Group.objects.get(name='Administrador')
                user.groups.add(admin_group)
            elif role == 'asistencia':
                # Asignar al grupo 'Asistencia'
                asistencia_group = Group.objects.get(name='Asistencia')
                user.groups.add(asistencia_group)

            # Mensaje de éxito
            messages.success(request, '¡Cuenta creada exitosamente! Ahora puedes iniciar sesión.')
            return redirect('login')  # Redirige al login después de crear la cuenta
        else:
            messages.error(request, 'Por favor, corrige los errores.')
    else:
        form = CustomUserCreationForm()

    return render(request, 'administrador/crear-administrador.html', {'form': form})